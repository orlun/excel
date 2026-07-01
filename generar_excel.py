#!/usr/bin/env python3
"""
Control Financiero Estudiantil — Versión Diaria
================================================
Flujo: Apuntas gastos cada día → Todo se actualiza solo
- Registro Diario (input principal)
- Dashboard (resumen + previsiones automáticas)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import range_boundaries

# ──────────────────────────────────────────────
# CONFIGURACIÓN
# ──────────────────────────────────────────────
CAPITAL = 6820
ALQUILER = 375
NUM_ROWS = 500  # filas para apuntar gastos

MESES_LABEL = [
    "Jul 2026","Ago 2026","Sep 2026","Oct 2026","Nov 2026","Dic 2026",
    "Ene 2027","Feb 2027","Mar 2027","Abr 2027","May 2027","Jun 2027"
]
MESES_NUM  = [7,8,9,10,11,12,1,2,3,4,5,6]
MESES_YEAR = [2026]*6 + [2027]*6

EUR = '#,##0.00 €'
PCT = '0.0%'
DATE_FMT = 'DD/MM/YYYY'

# Referencias al Registro Diario (rango de datos)
RD = "'Registro Diario'"
DR = f"{RD}!$B$5:$B${4+NUM_ROWS}"   # Fechas
DA = f"{RD}!$D$5:$D${4+NUM_ROWS}"   # Importes
DT = f"{RD}!$E$5:$E${4+NUM_ROWS}"   # Tipos

# ── PALETA ──
CN = "1A1F36"   # navy
CH = "252B48"   # header
CB = "4A6CF7"   # blue
CT = "06B6D4"   # teal
CG = "10B981"   # green
CR = "EF4444"   # red
CA = "F59E0B"   # amber
CW = "FFFFFF"
CBG= "F8FAFC"
CBD= "E2E8F0"
CRA= "F1F5F9"   # row alt
CCB= "EFF6FF"   # card blue
CCG= "ECFDF5"   # card green
CCR= "FEF2F2"   # card red
CCA= "FFFBEB"   # card amber
CCP= "F3E8FF"   # card purple

# ── ESTILOS ──
bd = Border(left=Side('thin',CBD), right=Side('thin',CBD),
            top=Side('thin',CBD), bottom=Side('thin',CBD))

def F(name='Calibri', sz=10, b=False, c="475569", u=None):
    return Font(name=name, size=sz, bold=b, color=c, underline=u)

ft_title  = F(sz=16, b=True, c=CW)
ft_sub    = F(sz=11, b=True, c=CW)
ft_sec    = F(sz=12, b=True, c=CN)
ft_hdr    = F(sz=10, b=True, c=CW)
ft_lbl    = F(sz=10, b=True, c="334155")
ft_n      = F()
ft_sm     = F(sz=9, c="94A3B8")
ft_kpi    = F(sz=16, b=True, c=CB)
ft_kpi_g  = F(sz=16, b=True, c=CG)
ft_kpi_r  = F(sz=16, b=True, c=CR)
ft_kpi_a  = F(sz=16, b=True, c=CA)
ft_vr     = F(b=True, c=CR)
ft_vg     = F(b=True, c=CG)
ft_vb     = F(sz=11, b=True, c=CB)
ft_input  = F(sz=11, c="1E293B")

def P(c): return PatternFill('solid', c)
fl_n  = P(CN); fl_h = P(CH); fl_b = P(CB); fl_w = P(CW)
fl_bg = P(CBG); fl_a = P(CRA)
fl_cb = P(CCB); fl_cg = P(CCG); fl_cr = P(CCR); fl_ca = P(CCA); fl_cp = P(CCP)
fl_rl = P("FADBD8"); fl_al = P("FEF9E7")
fl_input = P("F0F9FF")  # azul muy claro para celdas editables

ac = Alignment('center','center', wrap_text=True)
al = Alignment('left','center', wrap_text=True)
ar = Alignment('right','center')

def S(cell, font=None, fill=None, align=None, border=None, nf=None):
    if font: cell.font = font
    if fill: cell.fill = fill
    if align: cell.alignment = align
    if border: cell.border = border
    if nf: cell.number_format = nf

def merge(ws, rng, val, font, fill, align=ac):
    ws.merge_cells(rng)
    c = ws[rng.split(':')[0]]
    c.value = val; S(c, font, fill, align)
    mn_c, mn_r, mx_c, mx_r = range_boundaries(rng)
    for r in range(mn_r, mx_r+1):
        for cc in range(mn_c, mx_c+1):
            S(ws.cell(r, cc), fill=fill)

def kpi(ws, r, c, label, val, bg, ft=ft_kpi, nf=EUR, sp=2):
    ec = c + sp - 1
    for rr in range(r, r+3):
        for cc in range(c, ec+1):
            S(ws.cell(rr, cc), fill=bg, border=bd)
    merge(ws, f"{get_column_letter(c)}{r}:{get_column_letter(ec)}{r}", label, ft_sm, bg)
    merge(ws, f"{get_column_letter(c)}{r+1}:{get_column_letter(ec)}{r+1}", val, ft, bg)
    if nf and not isinstance(val, str):
        ws.cell(r+1, c).number_format = nf


# ══════════════════════════════════════════════
wb = openpyxl.Workbook()

# ╔═══════════════════════════════════════════╗
# ║  TAB 1: REGISTRO DIARIO                  ║
# ╚═══════════════════════════════════════════╝
wd = wb.active
wd.title = "Registro Diario"
wd.sheet_properties.tabColor = CT

widths_d = {'A':2, 'B':14, 'C':22, 'D':14, 'E':12, 'F':28, 'G':2}
for c, w in widths_d.items():
    wd.column_dimensions[c].width = w

# Fondo
last_row = NUM_ROWS + 10
for r in range(1, last_row):
    for c in range(1, 8):
        S(wd.cell(r, c), fill=fl_w)

# ── HEADER ──
for c in range(1, 8): S(wd.cell(1, c), fill=fl_n)
merge(wd, 'B2:F2', '✏️  REGISTRO DIARIO DE GASTOS', ft_title, fl_n)
for c in [1,7]: S(wd.cell(2, c), fill=fl_n)
wd.row_dimensions[2].height = 38

# Barra de estado rápida (auto-calculada)
for c in range(1, 8): S(wd.cell(3, c), fill=fl_h)
wd.row_dimensions[3].height = 30

# Saldo actual en la barra
wd.cell(3, 2).value = '💰 Saldo:'
S(wd.cell(3, 2), F(sz=10, b=True, c=CW), fl_h, al)
# Saldo = Capital - SUM gastos - alquiler pagado + ingresos
# Simplified: reference Dashboard calculation
wd.cell(3, 3).value = (
    f"={CAPITAL}"
    f"-SUMPRODUCT(({DT}=\"Gasto\")*{DA})"
    f"+SUMPRODUCT(({DT}=\"Ingreso\")*{DA})"
)
S(wd.cell(3, 3), F(sz=12, b=True, c=CG), fl_h, al, nf=EUR)

wd.cell(3, 4).value = '📊 Gastado:'
S(wd.cell(3, 4), F(sz=10, b=True, c=CW), fl_h, al)
wd.cell(3, 5).value = f"=SUMPRODUCT(({DT}=\"Gasto\")*{DA})"
S(wd.cell(3, 5), F(sz=12, b=True, c=CR), fl_h, al, nf=EUR)

# Acento
for c in range(1, 8): S(wd.cell(4, c), fill=fl_b)
wd.row_dimensions[4].height = 3

# ── CABECERAS TABLA ──
hdrs = ['Fecha', 'Concepto', 'Importe (€)', 'Tipo', 'Notas']
for i, h in enumerate(hdrs):
    cell = wd.cell(5, 2+i)
    cell.value = h
    S(cell, ft_hdr, fl_h, ac, bd)
wd.row_dimensions[5].height = 28
# Freeze panes para que las cabeceras queden fijas al hacer scroll
wd.freeze_panes = 'B6'

# ── DATA VALIDATION ──
dv_tipo = DataValidation(
    type="list", formula1='"Gasto,Ingreso"', allow_blank=True)
dv_tipo.error = "Selecciona Gasto o Ingreso"
dv_tipo.errorTitle = "Tipo inválido"
dv_tipo.prompt = "¿Es un Gasto o un Ingreso?"
dv_tipo.promptTitle = "Tipo"
wd.add_data_validation(dv_tipo)

# ── FILAS DE DATOS ──
for row in range(6, 6 + NUM_ROWS):
    bg = fl_a if (row - 6) % 2 == 0 else fl_w
    is_input = fl_input if (row - 6) % 2 == 0 else P("F8FDFF")
    
    # Fecha
    S(wd.cell(row, 2), ft_input, is_input, ac, bd, DATE_FMT)
    # Concepto
    S(wd.cell(row, 3), ft_input, is_input, al, bd)
    # Importe
    S(wd.cell(row, 4), ft_input, is_input, ac, bd, EUR)
    # Tipo
    cell_tipo = wd.cell(row, 5)
    S(cell_tipo, ft_input, is_input, ac, bd)
    dv_tipo.add(cell_tipo)
    # Notas
    S(wd.cell(row, 6), ft_sm, is_input, al, bd)

# ── TOTAL AL FINAL ──
total_row = 6 + NUM_ROWS
wd.cell(total_row, 2).value = '📊 TOTAL GASTOS'
S(wd.cell(total_row, 2), ft_lbl, fl_cb, al, bd)
wd.cell(total_row, 3).value = f'=COUNTIF(E6:E{total_row-1},"Gasto")&" movimientos"'
S(wd.cell(total_row, 3), ft_sm, fl_cb, ac, bd)
wd.cell(total_row, 4).value = f"=SUMPRODUCT(({DT}=\"Gasto\")*{DA})"
S(wd.cell(total_row, 4), ft_kpi_r, fl_cb, ac, bd, EUR)

total_row2 = total_row + 1
wd.cell(total_row2, 2).value = '📊 TOTAL INGRESOS'
S(wd.cell(total_row2, 2), ft_lbl, fl_cg, al, bd)
wd.cell(total_row2, 3).value = f'=COUNTIF(E6:E{total_row-1},"Ingreso")&" movimientos"'
S(wd.cell(total_row2, 3), ft_sm, fl_cg, ac, bd)
wd.cell(total_row2, 4).value = f"=SUMPRODUCT(({DT}=\"Ingreso\")*{DA})"
S(wd.cell(total_row2, 4), ft_kpi_g, fl_cg, ac, bd, EUR)

# ── INSTRUCCIONES ──
instr_row = total_row + 3
merge(wd, f'B{instr_row}:F{instr_row+1}',
    '📌 CÓMO USARLO: Pon la fecha, di qué compraste, cuánto costó, y elige "Gasto" o "Ingreso". '
    'El saldo de arriba se actualiza solo. Para alquiler: apúntalo como gasto con concepto "Alquiler". '
    'Ve a la pestaña "Dashboard" para ver el resumen completo y las previsiones.',
    F(sz=10, c="64748B"), fl_bg, al)

# Impresión
wd.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
wd.page_setup.fitToWidth = 1


# ╔═══════════════════════════════════════════╗
# ║  TAB 2: DASHBOARD                        ║
# ╚═══════════════════════════════════════════╝
ws = wb.create_sheet("Dashboard")
ws.sheet_properties.tabColor = CB

widths_ds = {'A':2,'B':17,'C':15,'D':15,'E':15,'F':15,'G':15,'H':15,'I':15,'J':2}
for c, w in widths_ds.items():
    ws.column_dimensions[c].width = w

for r in range(1, 85):
    for c in range(1, 11):
        S(ws.cell(r, c), fill=fl_w)

# ── HEADER ──
for c in range(1, 11): S(ws.cell(1, c), fill=fl_n)
merge(ws, 'B2:I2', '💰  DASHBOARD — Control Financiero Jul 2026 · Jun 2027', ft_title, fl_n)
for c in [1,10]: S(ws.cell(2, c), fill=fl_n)
merge(ws, 'B3:I3',
    f'Capital: {CAPITAL:,} €  ·  Alquiler: {ALQUILER} €/mes  ·  Todo se calcula del Registro Diario',
    F(sz=10, c="94A3B8"), fl_h)
for c in [1,10]: S(ws.cell(3, c), fill=fl_h)
ws.row_dimensions[2].height = 38
ws.row_dimensions[3].height = 22
for c in range(1, 11): S(ws.cell(4, c), fill=fl_b)
ws.row_dimensions[4].height = 3

# ── FÓRMULAS AUXILIARES (celdas ocultas para cálculos) ──
# Usamos columna J para helpers (no visible, fuera del rango visual)
# J1 = Total gastos sin alquiler (solo variable)
# J2 = Total ingresos
# J3 = Días transcurridos desde 1 Jul 2026
# J4 = Días restantes hasta 30 Jun 2027

# No, better to put helpers in visible cells with labels. Let me use a clean approach.
# I'll compute everything inline in the KPI formulas.

# ── KPIs PRINCIPALES ──
merge(ws, 'B6:I6', '📊  TU SITUACIÓN AHORA MISMO', ft_sec, fl_bg, al)
ws.row_dimensions[6].height = 26

# Saldo actual = Capital - todos los gastos + todos los ingresos
saldo_formula = f"={CAPITAL}-SUMPRODUCT(({DT}=\"Gasto\")*{DA})+SUMPRODUCT(({DT}=\"Ingreso\")*{DA})"
kpi(ws, 8, 2, '💰 SALDO ACTUAL', saldo_formula, fl_cg, ft_kpi_g)

# Total gastado
gasto_formula = f"=SUMPRODUCT(({DT}=\"Gasto\")*{DA})"
kpi(ws, 8, 4, '🔻 TOTAL GASTADO', gasto_formula, fl_cr, ft_kpi_r)

# Total ingresos extra
ingreso_formula = f"=SUMPRODUCT(({DT}=\"Ingreso\")*{DA})"
kpi(ws, 8, 6, '🔺 INGRESOS EXTRA', ingreso_formula, fl_cg, ft_kpi_g)

# Media diaria de gasto
media_formula = f'=IFERROR(SUMPRODUCT(({DT}="Gasto")*{DA})/MAX(1,TODAY()-DATE(2026,7,1)),0)'
kpi(ws, 8, 8, '📅 MEDIA GASTO/DÍA', media_formula, fl_ca, ft_kpi_a)

# ── PREVISIONES ──
merge(ws, 'B12:I12', '🔮  PREVISIONES (basadas en tu ritmo de gasto actual)', ft_sec, fl_bg, al)
ws.row_dimensions[12].height = 26

prev_headers = ['Indicador', 'Valor', 'Detalle']
for i, h in enumerate(prev_headers):
    cell = ws.cell(14, 2+i)
    cell.value = h
    S(cell, ft_hdr, fl_h, ac, bd)

# Fila 15: Presupuesto diario recomendado
# = (Saldo actual - alquiler restante) / días restantes
# Alquiler restante: meses que quedan * 375
ws.cell(15, 2).value = '💵 Presupuesto diario recomendado'
S(ws.cell(15, 2), ft_lbl, fl_cb, al, bd)
ws.cell(15, 3).value = (
    f'=IFERROR('
    f'({saldo_formula.lstrip("=")}'
    f'-MAX(0,12-MAX(1,MONTH(TODAY())-6+IF(MONTH(TODAY())<7,6,0)))*{ALQUILER})'
    f'/MAX(1,DATE(2027,6,30)-TODAY())'
    f',0)'
)
S(ws.cell(15, 3), ft_kpi, fl_cb, ac, bd, EUR)
ws.cell(15, 4).value = 'Lo que puedes gastar al día para llegar a junio 2027'
S(ws.cell(15, 4), ft_sm, fl_cb, al, bd)

# Fila 16: Gasto mensual proyectado
ws.cell(16, 2).value = '📊 Gasto mensual proyectado'
S(ws.cell(16, 2), ft_lbl, fl_a, al, bd)
ws.cell(16, 3).value = f'=IFERROR(SUMPRODUCT(({DT}="Gasto")*{DA})/MAX(1,TODAY()-DATE(2026,7,1))*30.44,0)'
S(ws.cell(16, 3), ft_kpi_a, fl_a, ac, bd, EUR)
ws.cell(16, 4).value = 'Proyección basada en tu media diaria × 30 días'
S(ws.cell(16, 4), ft_sm, fl_a, al, bd)

# Fila 17: Saldo previsto a fin de periodo
ws.cell(17, 2).value = '🎯 Saldo previsto a Jun 2027'
S(ws.cell(17, 2), ft_lbl, fl_cb, al, bd)
# = Saldo actual - (media diaria * días restantes)
ws.cell(17, 3).value = (
    f'=IFERROR('
    f'{saldo_formula.lstrip("=")}' 
    f'-(SUMPRODUCT(({DT}="Gasto")*{DA})/MAX(1,TODAY()-DATE(2026,7,1)))'
    f'*MAX(0,DATE(2027,6,30)-TODAY())'
    f',0)'
)
S(ws.cell(17, 3), ft_kpi, fl_cb, ac, bd, EUR)
ws.cell(17, 4).value = 'Si sigues gastando al mismo ritmo'
S(ws.cell(17, 4), ft_sm, fl_cb, al, bd)

# Fila 18: Fecha límite del dinero
ws.cell(18, 2).value = '⏰ Tu dinero dura hasta'
S(ws.cell(18, 2), ft_lbl, fl_a, al, bd)
# = HOY() + (Saldo actual / media diaria)
ws.cell(18, 3).value = (
    f'=IFERROR('
    f'TODAY()+({saldo_formula.lstrip("=")})'
    f'/(SUMPRODUCT(({DT}="Gasto")*{DA})/MAX(1,TODAY()-DATE(2026,7,1)))'
    f',"Sin datos aún")'
)
S(ws.cell(18, 3), F(sz=14, b=True, c=CB), fl_a, ac, bd, DATE_FMT)
ws.cell(18, 4).value = 'Fecha estimada en que se agota el capital a este ritmo'
S(ws.cell(18, 4), ft_sm, fl_a, al, bd)

# Fila 19: Estado
ws.cell(19, 2).value = '🚦 Estado'
S(ws.cell(19, 2), ft_lbl, fl_cg, al, bd)
ws.cell(19, 3).value = (
    f'=IFERROR(IF(({saldo_formula.lstrip("=")})/{CAPITAL}>0.5,'
    f'"🟢 Bien",'
    f'IF(({saldo_formula.lstrip("=")})/{CAPITAL}>0.2,'
    f'"🟡 Ojo",'
    f'IF(({saldo_formula.lstrip("=")})>0,'
    f'"🟠 Cuidado","🔴 Déficit"))),"🔵 Sin datos")'
)
S(ws.cell(19, 3), F(sz=14, b=True, c=CN), fl_cg, ac, bd)
ws.cell(19, 4).value = '🟢 >50% capital · 🟡 >20% · 🟠 <20% · 🔴 Déficit'
S(ws.cell(19, 4), ft_sm, fl_cg, al, bd)

# Formato condicional para saldo previsto
ws.conditional_formatting.add('C17',
    CellIsRule('lessThan', ['0'], fill=P("FADBD8"), font=F(sz=14, b=True, c=CR)))
ws.conditional_formatting.add('C17',
    CellIsRule('between', ['0','500'], fill=P("FEF9E7"), font=F(sz=14, b=True, c=CA)))


# ── RESUMEN MENSUAL AUTO ──
merge(ws, 'B22:I22', '📅  RESUMEN MENSUAL (automático desde tu Registro Diario)', 
      ft_sec, fl_bg, al)
ws.row_dimensions[22].height = 26

hdrs_m = ['Mes', 'Alquiler', 'Otros Gastos', 'Ingresos', 'Saldo Mes', 'Acumulado']
for i, h in enumerate(hdrs_m):
    cell = ws.cell(24, 2+i)
    cell.value = h
    S(cell, ft_hdr, fl_h, ac, bd)

for m in range(12):
    r = 25 + m
    mn = MESES_NUM[m]
    yr = MESES_YEAR[m]
    bg = fl_a if m % 2 == 0 else fl_w
    
    # Mes
    ws.cell(r, 2).value = MESES_LABEL[m]
    S(ws.cell(r, 2), ft_lbl, bg, ac, bd)
    
    # Alquiler: fijo editable
    ws.cell(r, 3).value = ALQUILER
    S(ws.cell(r, 3), ft_n, bg, ac, bd, EUR)
    
    # Otros Gastos (auto SUMPRODUCT desde Registro Diario)
    gastos_f = (
        f'=SUMPRODUCT(({DT}="Gasto")'
        f'*(MONTH({DR})={mn})*(YEAR({DR})={yr})'
        f'*{DA})'
    )
    ws.cell(r, 4).value = gastos_f
    S(ws.cell(r, 4), ft_vr, bg, ac, bd, EUR)
    
    # Ingresos (auto)
    ingresos_f = (
        f'=SUMPRODUCT(({DT}="Ingreso")'
        f'*(MONTH({DR})={mn})*(YEAR({DR})={yr})'
        f'*{DA})'
    )
    ws.cell(r, 5).value = ingresos_f
    S(ws.cell(r, 5), ft_vg, bg, ac, bd, EUR)
    
    # Saldo Mes = Ingresos - Alquiler - Gastos
    ws.cell(r, 6).value = f"=E{r}-C{r}-D{r}"
    S(ws.cell(r, 6), ft_vb, bg, ac, bd, EUR)
    
    # Acumulado = Capital + suma de saldos hasta aquí
    ws.cell(r, 7).value = f"={CAPITAL}+SUM($F$25:F{r})"
    S(ws.cell(r, 7), F(sz=11, b=True, c=CN), bg, ac, bd, EUR)

# Formato condicional acumulado
ws.conditional_formatting.add('G25:G36',
    CellIsRule('lessThan', ['500'], fill=fl_rl, font=F(b=True, c=CR)))
ws.conditional_formatting.add('G25:G36',
    CellIsRule('between', ['500','1500'], fill=fl_al, font=F(b=True, c=CA)))

# Data bar en acumulado
ws.conditional_formatting.add('G25:G36',
    DataBarRule(start_type='num', start_value=0, end_type='num', 
               end_value=CAPITAL, color=CB))

# Totales
tr = 37
ws.cell(tr, 2).value = '📊 TOTAL'
S(ws.cell(tr, 2), ft_lbl, fl_cb, ac, bd)
for ci in range(3, 8):
    cl = get_column_letter(ci)
    if ci <= 6:
        ws.cell(tr, ci).value = f"=SUM({cl}25:{cl}36)"
    elif ci == 7:
        ws.cell(tr, ci).value = "=G36"
    fnt = ft_vr if ci in (3,4) else ft_vg if ci == 5 else ft_kpi if ci == 7 else ft_vb
    S(ws.cell(tr, ci), fnt, fl_cb, ac, bd, EUR)

# Media
tr2 = 38
ws.cell(tr2, 2).value = '📊 MEDIA/MES'
S(ws.cell(tr2, 2), ft_lbl, fl_ca, ac, bd)
for ci in range(3, 7):
    cl = get_column_letter(ci)
    ws.cell(tr2, ci).value = f"={cl}{tr}/12"
    S(ws.cell(tr2, ci), ft_n, fl_ca, ac, bd, EUR)


# ── GRÁFICO 1: EVOLUCIÓN ACUMULADO ──
ch1 = LineChart()
ch1.title = "Evolución del Saldo Acumulado (€)"
ch1.style = 10; ch1.width = 30; ch1.height = 14
ch1.y_axis.title = "€"
d1 = Reference(ws, min_col=7, min_row=24, max_row=36)
c1 = Reference(ws, min_col=2, min_row=25, max_row=36)
ch1.add_data(d1, titles_from_data=True)
ch1.set_categories(c1)
ln = ch1.series[0]
ln.graphicalProperties.line.solidFill = CB
ln.graphicalProperties.line.width = 30000
ws.add_chart(ch1, "B40")

# ── GRÁFICO 2: GASTOS POR MES ──
ch2 = BarChart()
ch2.type = "col"; ch2.title = "Gastos por Mes (€)"
ch2.style = 10; ch2.width = 30; ch2.height = 14
ch2.y_axis.title = "€"
# Alquiler
d2a = Reference(ws, min_col=3, min_row=24, max_row=36)
ch2.add_data(d2a, titles_from_data=True)
# Otros gastos
d2b = Reference(ws, min_col=4, min_row=24, max_row=36)
ch2.add_data(d2b, titles_from_data=True)
c2 = Reference(ws, min_col=2, min_row=25, max_row=36)
ch2.set_categories(c2)
ch2.grouping = "stacked"
ws.add_chart(ch2, "B56")

# ── NOTA FINAL ──
merge(ws, 'B73:I73',
    '⚠️  TODO se calcula automáticamente desde tu "Registro Diario". '
    'Solo necesitas apuntar gastos allí cada día. Los alquileres mensuales de la tabla de arriba '
    'ya están a 375 €, pero puedes cambiarlos si algún mes varía.',
    ft_sm, fl_bg, al)


# ══════════════════════════════════════════════
# GUARDAR
# ══════════════════════════════════════════════
out = '/Users/ivansanchez/Desktop/personal/excel/Control_Financiero_2026.xlsx'
wb.save(out)

capital_tras_alq = CAPITAL - ALQUILER * 12
pres_dia = capital_tras_alq / 365

print(f"\n{'='*58}")
print(f"  ✅ Excel generado")
print(f"  📁 {out}")
print(f"{'='*58}")
print(f"\n  📊 2 Pestañas:")
print(f"     1. Registro Diario — Apunta gastos aquí cada día")
print(f"     2. Dashboard       — Todo automático: resumen,")
print(f"                          previsiones, gráficos")
print(f"\n  ✍️  Solo rellenas: Fecha + Concepto + Importe + Tipo")
print(f"     El resto se calcula solo.")
print(f"{'='*58}\n")
